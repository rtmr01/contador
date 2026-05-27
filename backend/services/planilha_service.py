"""Serviço para geração da planilha de contagem de tráfego em memória.
Adaptado da FASE 8 original para utilizar dados do banco SQLite + io.BytesIO.
"""

import io
from datetime import datetime, timedelta
from collections import defaultdict
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import models

VEHICLE_TYPES = [
    "Motos", "Automóveis", "Van", "Camionetas", "2CB", "3CB", "4CB", "2C", "3C", "4C",
    "2S1", "2S2", "2S3", "3S1", "3S2", "3S3", "3T4", "3T6", "3R6", "3Q4", "2C2", "2C3",
    "3C2", "3C3", "3D4",
]

_THIN = Side(style="thin", color="000000")
_MEDIUM = Side(style="medium", color="000000")
BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
FONT_TITLE = Font(name="Cambria", size=11, bold=True)
FONT_SUBTITLE = Font(name="Cambria", size=10, bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
FILL_HEADER = PatternFill(fill_type="solid", fgColor="EDEDED")

def _safe_str(v) -> str:
    if v is None: return ""
    return str(v).strip()

def _dia_semana(v) -> str:
    data = _safe_str(v)
    dias = ["SEGUNDA-FEIRA", "TERÇA-FEIRA", "QUARTA-FEIRA", "QUINTA-FEIRA", "SEXTA-FEIRA", "SÁBADO", "DOMINGO"]
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try: return dias[datetime.strptime(data, fmt).weekday()]
        except ValueError: continue
    return data

def _merge(ws, rng: str, value="", font=None, align=None, fill=None, border=True):
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]
    cell.value = value
    if font: cell.font = font
    if align: cell.alignment = align
    if fill: cell.fill = fill
    if border:
        for row in ws[rng]:
            for c in row: c.border = BORDER_THIN
    return cell

def _style_cell(cell, *, value=None, font=None, align=None, fill=None, border=True):
    if value is not None: cell.value = value
    if font: cell.font = font
    if align: cell.alignment = align
    if fill: cell.fill = fill
    if border: cell.border = BORDER_THIN

def _set_cell_border(cell, left=False, right=False, top=False, bottom=False):
    cell.border = Border(
        left=_THIN if left else None,
        right=_THIN if right else None,
        top=_THIN if top else None,
        bottom=_THIN if bottom else None,
    )

def _set_row_heights(ws):
    heights = {
        1: 11, 2: 20, 3: 16, 4: 30, 5: 20, 6: 20, 7: 20, 8: 20, 9: 20, 10: 20, 11: 20,
        12: 20, 13: 20, 14: 20, 15: 20, 16: 20, 17: 20, 18: 20, 19: 20, 20: 20, 21: 20,
        22: 20, 23: 20, 24: 20, 25: 20, 26: 20, 27: 20, 28: 20, 29: 20, 30: 20,
    }
    for r, h in heights.items(): ws.row_dimensions[r].height = h
    for col_idx in range(5, 31): ws.row_dimensions[col_idx].height = 20

def _set_col_widths(ws):
    widths = {"A": 2, "B": 3, "C": 27, "D": 23, "E": 12, "DV": 9}
    for col, w in widths.items(): ws.column_dimensions[col].width = w
    for col_idx in range(6, 126): ws.column_dimensions[get_column_letter(col_idx)].width = 7

def cria_cabecalho_legendas(ws, posto: str = "", rodovia: str = "", data: str = "", sentido: str = ""):
    _set_row_heights(ws)
    _set_col_widths(ws)

    _merge(ws, "B2:C2", "POSTO:", font=FONT_TITLE, align=ALIGN_CENTER, fill=FILL_HEADER)
    _style_cell(ws["D2"], value=_safe_str(posto), font=FONT_TITLE, align=ALIGN_CENTER, fill=FILL_HEADER)
    _style_cell(ws["E2"], value="RODOVIA:", font=FONT_TITLE, align=ALIGN_CENTER, fill=FILL_HEADER)
    _merge(ws, "F2:U2", _safe_str(rodovia), font=FONT_TITLE, align=ALIGN_CENTER, fill=FILL_HEADER)
    _style_cell(ws["Y2"], value="DATA:", font=FONT_TITLE, align=ALIGN_CENTER, fill=FILL_HEADER)
    _merge(ws, "AE2:AO2", _safe_str(data), font=FONT_TITLE, align=ALIGN_CENTER, fill=FILL_HEADER)
    _merge(ws, "AT2:BD2", _safe_str(data), font=FONT_TITLE, align=ALIGN_CENTER, fill=FILL_HEADER)
    _merge(ws, "BI2:BS2", "DIA DA SEMANA:", font=FONT_TITLE, align=ALIGN_CENTER, fill=FILL_HEADER)
    _merge(ws, "BX2:BZ2", _dia_semana(data), font=FONT_TITLE, align=ALIGN_CENTER, fill=FILL_HEADER)
    _merge(ws, "CW2:DV2", _safe_str(sentido), font=FONT_TITLE, align=ALIGN_LEFT, fill=FILL_HEADER)

    first_col = 2
    last_col = 125
    for col_idx in range(first_col, last_col + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = FILL_HEADER
        _set_cell_border(cell, left=(col_idx == first_col), right=(col_idx == last_col), top=True, bottom=True)

    _merge(ws, "B3:D4", "TIPOS DE VEÍCULOS", font=FONT_SUBTITLE, align=ALIGN_CENTER, fill=FILL_HEADER)
    _merge(ws, "E3:E4", "LEGENDA", font=FONT_SUBTITLE, align=ALIGN_CENTER, fill=FILL_HEADER)
    _merge(ws, "F3:DU3", "INTERVALO", font=FONT_SUBTITLE, align=ALIGN_CENTER, fill=FILL_HEADER)
    _merge(ws, "DV3:DV4", "TOTAL", font=FONT_SUBTITLE, align=ALIGN_CENTER, fill=FILL_HEADER)

    for hour in range(24):
        base_col = 6 + hour * 5
        next_hour = (hour + 1) % 24
        _style_cell(ws.cell(row=4, column=base_col), value=f"{hour:02d}:00  {next_hour:02d}:00", font=FONT_SUBTITLE, align=ALIGN_CENTER, fill=FILL_HEADER)
        intervalo_labels = [f"{hour:02d}:00  {hour:02d}:15", f"{hour:02d}:15 {hour:02d}:30", f"{hour:02d}:30 {hour:02d}:45", f"{hour:02d}:45  {next_hour:02d}:00"]
        for offset, label in enumerate(intervalo_labels, start=1):
            _style_cell(ws.cell(row=4, column=base_col + offset), value=label, font=FONT_SUBTITLE, align=ALIGN_CENTER, fill=FILL_HEADER)
            
    _merge(ws, "B5:C8", "PASSEIO E\nUTILITÁRIOS", font=FONT_SUBTITLE, align=Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90), fill=None)
    _merge(ws, "B9:C11", "ÔNIBUS", font=FONT_SUBTITLE, align=Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90), fill=None)
    _merge(ws, "B12:B29", "CAMINHÕES", font=FONT_SUBTITLE, align=Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90), fill=None)
    _merge(ws, "C12:C14", "LEVES", font=FONT_SUBTITLE, align=Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90), fill=None)
    _merge(ws, "C15:C20", "SEMI REBOQUES", font=FONT_SUBTITLE, align=Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90), fill=None)
    _merge(ws, "C21:C23", "SEMI\nREBOQUES\nESPECIAIS", font=FONT_SUBTITLE, align=Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90), fill=None)
    _merge(ws, "C24:C29", "REBOQUES", font=FONT_SUBTITLE, align=Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90), fill=None)
    _merge(ws, "B30:E30", "TOTAL", font=FONT_SUBTITLE, align=ALIGN_CENTER, fill=None)

    for offset, value in enumerate(VEHICLE_TYPES, start=5):
        cell = ws.cell(row=offset, column=5)
        _style_cell(cell, value=value, font=FONT_SUBTITLE, align=ALIGN_CENTER, fill=None)

    for row_idx in range(5, 30):
        for col_idx in range(6, 126):
            cell = ws.cell(row=row_idx, column=col_idx)
            _set_cell_border(cell, left=True, right=True, top=True, bottom=True)

    formula_start_cols = set(range(6, 122, 5))
    for row_idx in range(5, 30):
        for col_idx in range(6, 127):
            if col_idx in formula_start_cols: continue
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name=FONT_SUBTITLE.name, size=FONT_SUBTITLE.size, bold=False)
            cell.alignment = ALIGN_CENTER
            cell.value = "-"

    for row_idx in range(5, 30):
        for start_col in range(6, 122, 5):
            first_col = start_col + 1
            last_col = start_col + 4
            cell = ws.cell(row=row_idx, column=start_col)
            cell.value = f"=SUM({get_column_letter(first_col)}{row_idx}:{get_column_letter(last_col)}{row_idx})"
            cell.font = Font(name=FONT_SUBTITLE.name, size=FONT_SUBTITLE.size, bold=False)
            cell.alignment = ALIGN_CENTER

        total_ranges = []
        for start_col in range(6, 122, 5):
            first_col = start_col + 1
            last_col = start_col + 4
            total_ranges.append(f"{get_column_letter(first_col)}{row_idx}:{get_column_letter(last_col)}{row_idx}")
        total_formula = ",".join(total_ranges)
        total_cell = ws.cell(row=row_idx, column=126)
        total_cell.value = f"=SUM({total_formula})"
        total_cell.font = Font(name=FONT_SUBTITLE.name, size=FONT_SUBTITLE.size, bold=True)
        total_cell.alignment = ALIGN_CENTER
        _set_cell_border(total_cell, left=True, right=True, top=True, bottom=True)

    for row_idx in range(5, 30):
        cell = ws.cell(row=row_idx, column=4)
        _set_cell_border(cell, left=True, right=True, top=True, bottom=True)

    for col_idx in range(6, 127):
        cell = ws.cell(row=30, column=col_idx)
        col_letter = get_column_letter(col_idx)
        cell.value = f"=SUM({col_letter}5:{col_letter}29)"
        cell.font = Font(name=FONT_SUBTITLE.name, size=FONT_SUBTITLE.size, bold=True)
        cell.alignment = ALIGN_CENTER
        _set_cell_border(cell, left=True, right=True, top=True, bottom=True)


def generate_excel_from_counts(posto_name: str, rodovia: str, video_data_list: List[dict]) -> io.BytesIO:
    """
    Transforma listas de eventos do banco de dados em uma Planilha Excel completa (em memória).
    Itera sobre os vídeos fornecidos no video_data_list.
    """
    cat_map = {k.lower(): v for k, v in zip(
        ["moto", "automovel", "van", "camionetas", "2cb", "3cb", "4cb", "2c", "3c", "4c", "2s1", "2s2", "2s3", "3s1", "3s2", "3s3", "3t4", "3t6", "3r6", "3q4", "2c2", "2c3", "3c2", "3c3", "3d4", "carro"],
        VEHICLE_TYPES + ["Automóveis"]
    )}
    
    counts_dict = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int)))))
    
    first_video_start = datetime.now()
    if video_data_list and video_data_list[0].get("start_time"):
        first_video_start = video_data_list[0]["start_time"]
    
    for v_data in video_data_list:
        v_start = v_data.get("start_time") or datetime.now()
        counts = v_data.get("counts", [])
        
        for count in counts:
            dt = v_start + timedelta(seconds=count.video_timestamp)
            date_iso = dt.date().isoformat()
            hour = dt.hour
            slot = dt.minute // 15
            
            dir_str = str(count.direction).lower().strip()
            direction = 1
            if dir_str in ['2', '<-', 'volta', 'sul']: direction = 2
            
            cat_str = count.vehicle_type.lower().strip()
            cat = cat_map.get(cat_str)
            if not cat:
                if cat_str.startswith('auto') or cat_str == 'carro': cat = 'Automóveis'
                elif cat_str.startswith('moto'): cat = 'Motos'
                
            if cat:
                counts_dict[date_iso][direction][cat][hour][slot] += 1
                
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
        
    if not counts_dict:
        date_display = first_video_start.strftime('%d.%m')
        date_full = first_video_start.strftime('%d/%m/%Y')
        
        for direction in [1, 2]:
            nome_sentido = "IDA" if direction == 1 else "VOLTA"
            aba_nome = f"{date_display} - {nome_sentido}"
            
            ws = wb.create_sheet(title=aba_nome)
            cria_cabecalho_legendas(
                ws, 
                posto=posto_name, 
                rodovia=rodovia,
                data=date_full, 
                sentido=f"SENTIDO {direction} ({nome_sentido})"
            )
    else:
        for date_iso in sorted(counts_dict.keys()):
            dir_counts = counts_dict[date_iso]
            date_display = datetime.strptime(date_iso, '%Y-%m-%d').strftime('%d.%m')
            date_full = datetime.strptime(date_iso, '%Y-%m-%d').strftime('%d/%m/%Y')
            
            for direction in sorted(dir_counts.keys()):
                cat_counts = dir_counts[direction]
                nome_sentido = "IDA" if direction == 1 else "VOLTA"
                aba_nome = f"{date_display} - {nome_sentido}"
                
                ws = wb.create_sheet(title=aba_nome)
                cria_cabecalho_legendas(
                    ws, 
                    posto=posto_name, 
                    rodovia=rodovia,
                    data=date_full, 
                    sentido=f"SENTIDO {direction} ({nome_sentido})"
                )
                
                for cat, hour_counts in cat_counts.items():
                    row = 5 + VEHICLE_TYPES.index(cat)
                    for hour, slot_counts in hour_counts.items():
                        for slot, qtde in slot_counts.items():
                            col = 7 + hour * 5 + slot
                            cell = ws.cell(row=row, column=col)
                            cell.value = qtde

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
